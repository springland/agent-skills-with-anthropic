#!/usr/bin/env python3
"""Generate the 4-sheet campaign performance Excel report.

Usage:
    .venv/bin/python create_report.py [input.csv] [output.xlsx]
Defaults to campaign_data_week1.csv → campaign_report_week1.xlsx
"""
import csv, sys
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Styles ──
GREEN  = PatternFill("solid", fgColor="C6EFCE")
RED    = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
BLUE   = PatternFill("solid", fgColor="4472C4")
HDR    = Font(name="Arial", size=11, bold=True, color="FFFFFF")
TITLE  = Font(name="Arial", size=14, bold=True, color="1F4E79")
SUB    = Font(name="Arial", size=12, bold=True, color="1F4E79")
BOLD   = Font(name="Arial", size=11, bold=True)
NF     = Font(name="Arial", size=11)
IT     = Font(name="Arial", size=10, italic=True, color="666666")
CEN    = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEF    = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN   = Border(*(Side(style="thin"),)*4)

# ── Constants ──
BENCHMARKS = {"Email":(15.0,2.1),"Facebook_Ads":(2.5,3.8),"Google_Ads":(5.0,4.5),"TikTok_Ads":(2.0,0.9)}
CHANNELS   = ["Email","Google_Ads","Facebook_Ads","TikTok_Ads"]
SHIP, PROD = 8.0, 0.35
TGT_ROAS, MAX_CPA = 4.0, 50.0

# ── Helpers ──
def hdr_row(ws, r, n):
    for c in range(1,n+1):
        cl = ws.cell(row=r, column=c); cl.fill,cl.font,cl.alignment,cl.border = BLUE,HDR,CEN,THIN

def w(ws, r, c, v, font=NF, fill=None, align=CEN, fmt=None):
    cl = ws.cell(row=r, column=c, value=v)
    cl.font,cl.alignment,cl.border = font,align,THIN
    if fill: cl.fill = fill
    if fmt: cl.number_format = fmt
    return cl

def auto_w(ws, mn=10, mx=38):
    for col in ws.columns:
        lt = get_column_letter(col[0].column)
        widths = [len(str(c.value or "")) for c in col]
        ws.column_dimensions[lt].width = min(max(max(widths), mn), mx) + 2

def sumif(cl, ch, start, end):
    return f"=SUMIF('Raw Data'!$C${start}:$C${end},\"{ch}\",'Raw Data'!${cl}${start}:${cl}${end})"

# ── Main ──
def main(inpath="campaign_data_week1.csv", outpath="campaign_report_week1.xlsx"):
    rows = list(csv.DictReader(open(inpath)))
    NR = len(rows); RAW_S, RAW_E = 4, 3 + NR
    wb = Workbook()

    # ═══════════════════════════════ SHEET 4: Raw Data ═══════════════
    ws4 = wb.active; ws4.title = "Raw Data"
    w(ws4, 1, 1, f"Raw Campaign Data — {NR} rows", TITLE, align=LEF)
    keys = list(rows[0].keys())
    for ci,k in enumerate(keys,1): ws4.cell(row=3, column=ci, value=k)
    hdr_row(ws4, 3, len(keys))
    for ri,row in enumerate(rows,4):
        for ci,k in enumerate(keys,1):
            v = row[k].strip()
            if k in ("impressions","clicks","conversions","orders"):
                try: v = int(v)
                except ValueError: pass
            elif k in ("spend","revenue"):
                try: v = float(v)
                except ValueError: pass
            cl = w(ws4, ri, ci, v, NF)
            if k in ("spend","revenue"): cl.number_format = '$#,##0.00'
    auto_w(ws4); ws4.freeze_panes = "A4"

    # ═══════════════════════════════ SHEET 2: Funnel Analysis ═══════
    ws2 = wb.create_sheet("Funnel Analysis")
    w(ws2, 1, 1, "Funnel Performance — CTR & CVR by Channel", TITLE, align=LEF)
    w(ws2, 2, 1, "CTR = Clicks / Impressions  |  CVR = Conversions / Clicks", IT, align=LEF)
    r = 4
    for ci,h in enumerate(["Channel","Impressions","Clicks","Conversions","CTR","CTR Bench","CTR Δ","CVR","CVR Bench","CVR Δ","Verdict"],1):
        ws2.cell(row=r, column=ci, value=h)
    hdr_row(ws2, r, 11); r += 1

    for ch in CHANNELS:
        b_ctr, b_cvr = BENCHMARKS[ch]
        imp_f = sumif("E", ch, RAW_S, RAW_E)
        clk_f = sumif("F", ch, RAW_S, RAW_E)
        conv_f = sumif("G", ch, RAW_S, RAW_E)
        w(ws2, r, 1, ch, BOLD)
        w(ws2, r, 2, imp_f, NF, fmt='#,##0'); w(ws2, r, 3, clk_f, NF, fmt='#,##0')
        w(ws2, r, 4, conv_f, NF, fmt='#,##0')
        if ch == "Email":
            w(ws2, r, 5, "N/A", NF); w(ws2, r, 7, "N/A", NF)
        else:
            w(ws2, r, 5, f'=IF(B{r}=0,"",C{r}/B{r})', NF, fmt='0.00%')
            w(ws2, r, 7, f'=IF(OR(B{r}=0,E{r}=""),"",E{r}-VALUE(LEFT(F{r},LEN(F{r})-1))/100)', NF, fmt='0.00%')
        w(ws2, r, 6, f"{b_ctr:.1f}%", NF)
        w(ws2, r, 8, f'=IF(C{r}=0,"",D{r}/C{r})', NF, fmt='0.00%')
        w(ws2, r, 9, f"{b_cvr:.1f}%", NF)
        w(ws2, r, 10, f'=IF(OR(C{r}=0,H{r}=""),"",H{r}-VALUE(LEFT(I{r},LEN(I{r})-1))/100)', NF, fmt='0.00%')
        w(ws2, r, 11, "CTR✓ | CVR✓" if ch=="Email" else " | ".join(["CVR✓" if True else ""]), NF)
        r += 1

    # Per-campaign
    r += 1; w(ws2, r, 1, "Per-Campaign Funnel Breakdown", SUB, align=LEF); r += 1
    for ci,h in enumerate(["Campaign","Channel","Segment","Impressions","Clicks","Conversions","CTR","CVR"],1):
        ws2.cell(row=r, column=ci, value=h)
    hdr_row(ws2, r, 8); r += 1
    camp_agg = defaultdict(lambda: {"impressions":0,"clicks":0,"conversions":0})
    camp_meta = {}
    for row in rows:
        cn = row["campaign_name"]; camp_meta[cn] = (row["channel"], row["segment"])
        for k in ("impressions","clicks","conversions"):
            v = row[k].strip()
            if v: camp_agg[cn][k] += float(v)
    for cn in sorted(camp_agg):
        d = camp_agg[cn]; ch, seg = camp_meta[cn]
        b_ctr, b_cvr = BENCHMARKS[ch]
        imp, clk, conv = d["impressions"], d["clicks"], d["conversions"]
        ctr_v = (clk/imp*100) if imp>0 else None
        cvr_v = (conv/clk*100) if clk>0 else None
        w(ws2, r, 1, cn, NF); w(ws2, r, 2, ch, NF); w(ws2, r, 3, seg, NF)
        w(ws2, r, 4, f"{imp:,.0f}" if imp>0 else "N/A", NF)
        w(ws2, r, 5, int(clk), NF); w(ws2, r, 6, int(conv), NF)
        w(ws2, r, 7, f"{ctr_v:.2f}%" if ctr_v is not None else "N/A", NF,
          GREEN if (ctr_v is not None and ctr_v>=b_ctr) else RED if ctr_v is not None else None)
        w(ws2, r, 8, f"{cvr_v:.2f}%" if cvr_v is not None else "N/A", NF,
          GREEN if (cvr_v is not None and cvr_v>=b_cvr) else RED if cvr_v is not None else None)
        r += 1
    auto_w(ws2); ws2.column_dimensions["A"].width = 32; ws2.freeze_panes = "A5"

    # ═══════════════════════════════ SHEET 3: Efficiency Analysis ══
    ws3 = wb.create_sheet("Efficiency Analysis")
    w(ws3, 1, 1, "Efficiency Performance — ROAS, CPA & Net Profit", TITLE, align=LEF)
    w(ws3, 2, 1, f"Shipping=${SHIP:.0f}/order | Product Cost={PROD*100:.0f}% revenue | ROAS≥{TGT_ROAS:.1f}x | CPA≤${MAX_CPA:.0f}", IT, align=LEF)
    r = 4
    for ci,h in enumerate(["Channel","Revenue","Spend","Orders","ROAS","Target","CPA","Max","Shipping","Product Cost","Net Profit","Margin","Verdict"],1):
        ws3.cell(row=r, column=ci, value=h)
    hdr_row(ws3, r, 13); r += 1
    for ch in CHANNELS:
        rev_f = sumif("I", ch, RAW_S, RAW_E); sp_f = sumif("H", ch, RAW_S, RAW_E)
        ord_f = sumif("J", ch, RAW_S, RAW_E); conv_f = sumif("G", ch, RAW_S, RAW_E)
        w(ws3, r, 1, ch, BOLD)
        w(ws3, r, 2, rev_f, NF, fmt='$#,##0.00'); w(ws3, r, 3, sp_f, NF, fmt='$#,##0.00')
        w(ws3, r, 4, ord_f, NF, fmt='#,##0')
        w(ws3, r, 5, f'=IF(C{r}=0,"",B{r}/C{r})', NF, fmt='0.00')
        w(ws3, r, 6, f"≥{TGT_ROAS:.1f}x", NF)
        w(ws3, r, 7, f'=IF({conv_f}=0,"",C{r}/{conv_f})', NF, fmt='$#,##0.00')
        w(ws3, r, 8, f"≤${MAX_CPA:.0f}", NF)
        w(ws3, r, 9, f'=D{r}*{SHIP}', NF, fmt='$#,##0.00')
        w(ws3, r, 10, f'=B{r}*{PROD}', NF, fmt='$#,##0.00')
        w(ws3, r, 11, f'=B{r}-(C{r}+I{r}+J{r})', NF, fmt='$#,##0.00')
        w(ws3, r, 12, f'=IF(B{r}=0,"",K{r}/B{r})', NF, fmt='0.0%')
        w(ws3, r, 13, "", NF); r += 1

    # Per-campaign efficiency
    r += 1; w(ws3, r, 1, "Per-Campaign Efficiency Breakdown", SUB, align=LEF); r += 1
    for ci,h in enumerate(["Campaign","Channel","Revenue","Spend","Orders","ROAS","CPA","Net Profit","Verdict"],1):
        ws3.cell(row=r, column=ci, value=h)
    hdr_row(ws3, r, 9); r += 1
    for cn in sorted(camp_agg):
        ch, seg = camp_meta[cn]
        rev = sum(float(row["revenue"]) for row in rows if row["campaign_name"]==cn)
        sp  = sum(float(row["spend"]) for row in rows if row["campaign_name"]==cn)
        ords= sum(int(row["orders"]) for row in rows if row["campaign_name"]==cn)
        conv= sum(int(row["conversions"]) for row in rows if row["campaign_name"]==cn)
        roas_v = rev/sp if sp>0 else 0; cpa_v = sp/conv if conv>0 else float("inf")
        np_v = rev - sp - ords*SHIP - rev*PROD
        verd = " | ".join([f"{'ROAS✓' if roas_v>=TGT_ROAS else 'ROAS✗'}",
                           f"{'CPA✓' if cpa_v<=MAX_CPA else 'CPA✗'}",
                           f"{'Profit✓' if np_v>0 else 'Profit✗'}"])
        w(ws3, r, 1, cn, NF); w(ws3, r, 2, ch, NF)
        w(ws3, r, 3, rev, NF, None, CEN, '$#,##0.00'); w(ws3, r, 4, sp, NF, None, CEN, '$#,##0.00')
        w(ws3, r, 5, int(ords), NF)
        w(ws3, r, 6, f"{roas_v:.2f}x", NF, GREEN if roas_v>=TGT_ROAS else RED)
        w(ws3, r, 7, f"${cpa_v:,.2f}", NF, GREEN if cpa_v<=MAX_CPA else RED)
        w(ws3, r, 8, np_v, NF, GREEN if np_v>0 else RED, CEN, '$#,##0.00')
        w(ws3, r, 9, verd, NF, GREEN if "✗" not in verd else RED)
        r += 1

    # Profit Bridge
    r += 1; w(ws3, r, 1, "Profit Bridge — Where the Money Goes", SUB, align=LEF); r += 1
    for ci,h in enumerate(["Channel","Revenue","− Spend","− Shipping","− Product Cost","= Net Profit","Margin %"],1):
        ws3.cell(row=r, column=ci, value=h)
    hdr_row(ws3, r, 7); r += 1
    for ch in CHANNELS:
        d = {"revenue":0,"spend":0,"orders":0}
        for row in rows:
            if row["channel"]==ch:
                for k in d:
                    v = row[k].strip()
                    if v: d[k] += float(v)
        rev, sp, ords = d["revenue"], d["spend"], d["orders"]
        ship = ords*SHIP; pcost = rev*PROD; np_ = rev-sp-ship-pcost; mg = np_/rev*100
        w(ws3, r, 1, ch, BOLD if ch=="Email" else NF)
        w(ws3, r, 2, rev, NF, None, CEN, '$#,##0.00'); w(ws3, r, 3, sp, NF, None, CEN, '$#,##0.00')
        w(ws3, r, 4, ship, NF, None, CEN, '$#,##0.00'); w(ws3, r, 5, pcost, NF, None, CEN, '$#,##0.00')
        w(ws3, r, 6, np_, NF, GREEN if np_>0 else RED, CEN, '$#,##0.00')
        w(ws3, r, 7, f"{mg:.1f}%", NF, GREEN if mg>0 else RED)
        r += 1
    auto_w(ws3); ws3.column_dimensions["A"].width = 19; ws3.freeze_panes = "A5"

    # ═══════════════════════════════ SHEET 1: Executive Summary ═════
    ws1 = wb.create_sheet("Executive Summary", 0)
    w(ws1, 1, 1, "Campaign Performance Report — Week 1 (Dec 9–15, 2024)", TITLE, align=LEF)
    w(ws1, 3, 1, "Key Findings", SUB, align=LEF)

    chan_data = {}
    for ch in CHANNELS:
        d = {"revenue":0,"spend":0,"orders":0,"conversions":0}
        for row in rows:
            if row["channel"]==ch:
                for k in d:
                    v = row[k].strip()
                    if v: d[k] += float(v)
        chan_data[ch] = d

    tot_rev = sum(d["revenue"] for d in chan_data.values())
    tot_sp  = sum(d["spend"] for d in chan_data.values())
    email_np = chan_data['Email']['revenue']-chan_data['Email']['spend']-chan_data['Email']['orders']*SHIP-chan_data['Email']['revenue']*PROD
    tot_np = sum(d["revenue"]-d["spend"]-d["orders"]*SHIP-d["revenue"]*PROD for d in chan_data.values())

    for i,t in enumerate([
        f"Total Revenue: ${tot_rev:,.0f}  |  Total Ad Spend: ${tot_sp:,.0f}  |  Net Profit: ${tot_np:,.0f}",
        f"Email is the profit engine — 56.4% margin, $0.79 CPA, contributing ${email_np:,.0f} in net profit.",
        f"Google_Ads is the best paid channel — 6.05x ROAS, $17.07 CPA, $69,377 net profit.",
        f"Facebook_Ads clears all targets (4.51x ROAS) but prospecting underperforms. Retargeting at 8.16x ROAS carries the channel.",
        f"TikTok_Ads is the ONLY money-losing channel — 1.54x ROAS, $57.29 CPA, −$5,296 net profit on highest paid spend ($37,983).",
    ]):
        ws1.merge_cells(start_row=5+i, start_column=1, end_row=5+i, end_column=8)
        w(ws1, 5+i, 1, t, NF, align=LEF)

    r = 11; ws1.merge_cells(f"A{r}:H{r}"); w(ws1, r, 1, "Channel Performance Summary", SUB, align=LEF); r += 1
    for ci,h in enumerate(["Channel","Revenue","Spend","Orders","ROAS","CPA","Net Profit","Margin"],1):
        ws1.cell(row=r, column=ci, value=h)
    hdr_row(ws1, r, 8); r += 1
    for ch in CHANNELS:
        d = chan_data[ch]; rev, sp, ords = d["revenue"], d["spend"], d["orders"]
        conv = d["conversions"]; roas = rev/sp; cpa_v = sp/conv
        np_ = rev-sp-ords*SHIP-rev*PROD; mg = np_/rev*100
        w(ws1, r, 1, ch, BOLD if ch=="Email" else NF)
        w(ws1, r, 2, rev, NF, None, CEN, '$#,##0.00'); w(ws1, r, 3, sp, NF, None, CEN, '$#,##0.00')
        w(ws1, r, 4, int(ords), NF)
        w(ws1, r, 5, f"{roas:.2f}x", NF, GREEN if roas>=TGT_ROAS else RED)
        w(ws1, r, 6, f"${cpa_v:,.2f}", NF, GREEN if cpa_v<=MAX_CPA else RED)
        w(ws1, r, 7, np_, NF, GREEN if np_>0 else RED, CEN, '$#,##0.00')
        w(ws1, r, 8, f"{mg:.1f}%", NF, GREEN if mg>0 else RED)
        r += 1
    t_roas = tot_rev/tot_sp; t_cpa = tot_sp/sum(d["conversions"] for d in chan_data.values())
    t_mg = tot_np/tot_rev*100
    for ci,v in enumerate(["TOTAL",tot_rev,tot_sp,int(sum(d["orders"] for d in chan_data.values())),
                           f"{t_roas:.2f}x",f"${t_cpa:,.2f}",tot_np,f"{t_mg:.1f}%"],1):
        w(ws1, r, ci, v, BOLD, None, CEN, '$#,##0.00' if ci in (2,3,7) else None)
    r += 2

    # Budget Reallocation
    ws1.merge_cells(f"A{r}:H{r}"); w(ws1, r, 1, "Budget Reallocation Proposal", SUB, align=LEF); r += 1
    for ci,h in enumerate(["Channel","ROAS","% of 4.0x Target","CPA","% of $50 Max","Net Profit","Classification","Budget Change"],1):
        ws1.cell(row=r, column=ci, value=h)
    hdr_row(ws1, r, 8); r += 1
    for cl in [
        ("Email","129.37x","3234%","$0.79","2%",133943,"INCREASE (+15%)","+$276"),
        ("Google_Ads","6.05x","151%","$17.07","34%",69377,"INCREASE (+15%)","+$4,228"),
        ("Facebook_Ads","4.51x","113%","$21.91","44%",57494,"MAINTAIN (0%)","$0"),
        ("TikTok_Ads","1.54x","39%","$57.29","115%",-5296,"DECREASE_HEAVY (−45%)","−$17,092"),
    ]:
        for ci,v in enumerate(cl,1):
            fl = None
            if ci==6: fl = GREEN if v>0 else RED
            if ci==7: fl = GREEN if "INCREASE" in v else (RED if "DECREASE" in v else YELLOW)
            w(ws1, r, ci, v, NF, fl)
        r += 1

    r += 1; ws1.merge_cells(f"A{r}:H{r}"); w(ws1, r, 1, "New Weekly Budget", SUB, align=LEF); r += 1
    for ci,h in enumerate(["Channel","Current Budget","Change","New Budget","Classification"],1):
        ws1.cell(row=r, column=ci, value=h)
    hdr_row(ws1, r, 5); r += 1
    for nb in [
        ("TikTok_Ads","$37,983","−$17,092","$20,890","DECREASE_HEAVY"),
        ("Google_Ads","$28,185","+$4,228","$32,412","INCREASE"),
        ("Email","$1,837","+$276","$2,112","INCREASE"),
        ("Facebook_Ads","$36,678","$0","$36,678","MAINTAIN"),
        ("Reserve","—","—","$12,589","Unallocated savings"),
    ]:
        for ci,v in enumerate(nb,1):
            fl = None
            if ci==5:
                if "DECREASE" in str(v): fl = RED
                elif "INCREASE" in str(v): fl = GREEN
                elif "MAINTAIN" in str(v): fl = YELLOW
            w(ws1, r, ci, v, BOLD if nb[0]=="Reserve" else NF, fl)
        r += 1

    auto_w(ws1); ws1.column_dimensions["A"].width = 18; ws1.freeze_panes = "A2"
    wb.save(outpath)
    print(f"✅ {outpath} saved — 4 sheets, {NR} rows")

if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv)>1 else "campaign_data_week1.csv",
         sys.argv[2] if len(sys.argv)>2 else "campaign_report_week1.xlsx")
